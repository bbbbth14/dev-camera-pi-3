#!/usr/bin/env python3
"""
View attendance Excel file structure and data
Displays all sheets and their contents
"""

import os
from openpyxl import load_workbook
from datetime import datetime

def view_excel_file(file_path='data/attendance.xlsx'):
    """Display Excel file structure and contents"""
    
    if not os.path.exists(file_path):
        print(f"[ERROR] File not found: {file_path}")
        return
    
    print("=" * 80)
    print(f"ATTENDANCE FILE VIEWER: {file_path}")
    print("=" * 80)
    
    wb = load_workbook(file_path)
    
    print(f"\n📊 Total Sheets: {len(wb.sheetnames)}\n")
    
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        
        print("=" * 80)
        print(f"📄 Sheet: {sheet_name}")
        print("=" * 80)
        
        if sheet_name == 'Template':
            print("   (Hidden template sheet - used for creating new user sheets)")
            print()
            continue
        
        # Determine sheet type based on name
        if '_' in sheet_name and any(month in sheet_name for month in 
            ['January', 'February', 'March', 'April', 'May', 'June', 
             'July', 'August', 'September', 'October', 'November', 'December']):
            # New format - user monthly sheet
            view_monthly_sheet(ws, sheet_name)
        else:
            # Old format or other sheets
            view_generic_sheet(ws)
        
        print()

    wb.close()
    print("=" * 80)

def view_monthly_sheet(ws, sheet_name):
    """View a monthly per-user sheet (new format)"""
    
    # Extract user name and month
    parts = sheet_name.split('_')
    user = parts[0] if parts else "Unknown"
    
    # Show header
    title = ws['A1'].value
    print(f"   User: {user}")
    print(f"   Title: {title}")
    print()
    
    # Show column headers
    headers = [cell.value for cell in ws[2]]
    print("   Columns:", " | ".join(headers))
    print("   " + "-" * 75)
    
    # Show attendance records (only rows with data)
    today = datetime.now().strftime('%Y-%m-%d')
    record_count = 0
    
    for row in ws.iter_rows(min_row=3, values_only=True):
        date_val = row[0]
        first_in = row[2] if len(row) > 2 else None
        last_out = row[3] if len(row) > 3 else None
        
        # Only show rows with attendance data
        if first_in or last_out:
            record_count += 1
            
            # Format row
            day = row[1] if len(row) > 1 else ''
            total = row[4] if len(row) > 4 else ''
            status = row[5] if len(row) > 5 else ''
            late = row[6] if len(row) > 6 else ''
            ot = row[7] if len(row) > 7 else ''
            
            # Highlight today
            marker = " 👈 TODAY" if date_val == today else ""
            
            print(f"   {date_val} ({day:3s}) | In: {first_in or '---':8s} | Out: {last_out or '---':8s} | "
                  f"Total: {total or '---':7s} | {status or '---':8s} | Late: {late or '---':4s} | "
                  f"OT: {ot or '---':4s}{marker}")
    
    if record_count == 0:
        print("   (No attendance recorded this month)")
    else:
        print()
        print(f"   📈 Total days with attendance: {record_count}")

def view_generic_sheet(ws):
    """View a generic sheet (old format or summary)"""
    
    print(f"   Total rows: {ws.max_row}")
    print(f"   Total columns: {ws.max_column}")
    print()
    
    # Show first row (usually headers)
    print("   Headers:")
    headers = [cell.value for cell in ws[1]]
    print("   ", " | ".join(str(h) for h in headers if h))
    print()
    
    # Show a few data rows
    print("   Sample data (first 10 rows):")
    for idx, row in enumerate(ws.iter_rows(min_row=2, max_row=11, values_only=True), start=2):
        if any(row):  # Only show non-empty rows
            row_str = " | ".join(str(cell) if cell else "---" for cell in row)
            print(f"   Row {idx}: {row_str}")
    
    if ws.max_row > 11:
        print(f"   ... ({ws.max_row - 11} more rows)")

if __name__ == '__main__':
    import sys
    file_path = sys.argv[1] if len(sys.argv) > 1 else 'data/attendance.xlsx'
    view_excel_file(file_path)
