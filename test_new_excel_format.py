#!/usr/bin/env python3
"""
Test script for new monthly Excel format
Tests the one-sheet-per-user format with monthly calendar
"""

import os
import sys
from datetime import datetime
from attendance_tracker import AttendanceTracker

def test_new_format():
    """Test the new Excel format with multiple users"""
    
    print("=" * 60)
    print("Testing New Monthly Excel Format")
    print("=" * 60)
    
    # Create tracker
    tracker = AttendanceTracker()
    
    # Test with multiple users
    test_users = ["Linh", "John", "Mary"]
    
    print("\n" + "=" * 60)
    print("TEST 1: Recording CHECK-INs for multiple users")
    print("=" * 60)
    
    for user in test_users:
        result = tracker.record_event(user, 'CHECK_IN')
        print(f"✓ {user} checked in: {result}")
    
    print("\n" + "=" * 60)
    print("TEST 2: Getting user status (should all be IN)")
    print("=" * 60)
    
    status = tracker.get_user_status()
    for user, info in status.items():
        print(f"{user}: Status={info['status']}, Time={info['check_in_time']}, Late={info.get('time_late', 'N/A')}")
    
    print("\n" + "=" * 60)
    print("TEST 3: Recording CHECK-OUTs for users")
    print("=" * 60)
    
    for user in test_users[:2]:  # Only check out first 2 users
        result = tracker.record_event(user, 'CHECK_OUT')
        print(f"✓ {user} checked out: {result}")
    
    print("\n" + "=" * 60)
    print("TEST 4: Getting updated user status")
    print("=" * 60)
    
    status = tracker.get_user_status()
    for user, info in status.items():
        print(f"{user}: Status={info['status']}, Total={info.get('duration', 'N/A')}, OT={info.get('time_ot', '0m')}")
    
    print("\n" + "=" * 60)
    print("TEST 5: Getting today's attendance summary")
    print("=" * 60)
    
    records = tracker.get_today_attendance()
    for record in records:
        print(f"{record['Name']}: In={record['Time In']}, Out={record['Time Out']}, Total={record['Total']}")
    
    print("\n" + "=" * 60)
    print("TEST 6: Verifying Excel file structure")
    print("=" * 60)
    
    from openpyxl import load_workbook
    test_file = tracker.attendance_file
    wb = load_workbook(test_file)
    
    print(f"Total sheets: {len(wb.sheetnames)}")
    for sheet_name in wb.sheetnames:
        ws = wb[sheet_name]
        print(f"\nSheet: {sheet_name}")
        
        if sheet_name == 'Template':
            print("  (Hidden template sheet)")
            continue
        
        # Print header
        print(f"  Title: {ws['A1'].value}")
        
        # Print column headers
        headers = [cell.value for cell in ws[2]]
        print(f"  Headers: {headers}")
        
        # Count filled rows (today's attendance)
        filled_rows = 0
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if row[2].value or row[3].value:  # Has check-in or check-out
                filled_rows += 1
                print(f"  Row {row[0].row}: Date={row[0].value}, In={row[2].value}, Out={row[3].value}, Total={row[4].value}")
        
        if filled_rows == 0:
            print("  (No attendance recorded today)")
    
    wb.close()
    
    print("\n" + "=" * 60)
    print("✓ All tests completed!")
    print(f"✓ Test file created: {test_file}")
    print("=" * 60)
    print("\nNew format features:")
    print("  • One sheet per user (e.g., 'Linh_December_2024')")
    print("  • Pre-filled with all days of the month")
    print("  • Tracks First In, Last Out, Total Hours, Late time, Overtime")
    print("  • Weekends highlighted in gray")
    print("  • Monthly view for easy time tracking")
    print("=" * 60)

if __name__ == '__main__':
    test_new_format()
