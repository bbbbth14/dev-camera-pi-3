"""
Attendance/Check-in/Check-out Tracker
Manages attendance records and access control
"""

import csv
import os
import threading
import calendar
import uuid
import hashlib
from datetime import datetime, timedelta
from typing import Optional, List, Dict
from zipfile import BadZipFile
import config
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[WARNING] openpyxl not installed. Run: pip install openpyxl")


class AttendanceTracker:
    """Track check-in/check-out events"""
    
    def __init__(self):
        """Initialize the attendance tracker"""
        self.attendance_file = config.ATTENDANCE_FILE
        self.status_log_file = os.path.join(config.DATA_DIR, 'status_log.csv')
        self.user_ids_file = os.path.join(config.DATA_DIR, 'user_ids.csv')
        self.last_checkin = {}  # Track last check-in time per person
        self.user_ids = {}  # Map user names to IDs
        self._excel_lock = threading.Lock()
        
        # Load user IDs
        self._load_user_ids()
        
        # Create (or recover) attendance file
        if not os.path.exists(self.attendance_file):
            self._create_attendance_file()
        else:
            # If the XLSX is corrupted (common if read/write happens concurrently),
            # rename it and recreate a fresh workbook so the app can continue.
            if EXCEL_AVAILABLE:
                try:
                    with self._excel_lock:
                        wb = load_workbook(self.attendance_file)
                        wb.close()
                except BadZipFile:
                    ts = datetime.now().strftime('%Y%m%d_%H%M%S')
                    corrupt_path = self.attendance_file.replace('.xlsx', f'_corrupt_{ts}.xlsx')
                    try:
                        os.rename(self.attendance_file, corrupt_path)
                        print(f"[WARNING] Attendance file corrupted. Renamed to: {corrupt_path}")
                    except Exception as e:
                        print(f"[WARNING] Failed to rename corrupted attendance file: {e}")
                    self._create_attendance_file()
        
        # Create status log file if it doesn't exist
        if not os.path.exists(self.status_log_file):
            self._create_status_log_file()
        
        # Load last check-ins from today
        self._load_today_checkins()
        
        print("[INFO] Attendance tracker initialized")
    
    def _load_user_ids(self):
        """Load user IDs from file or create new file"""
        try:
            if os.path.exists(self.user_ids_file):
                with open(self.user_ids_file, 'r') as f:
                    reader = csv.reader(f)
                    next(reader)  # Skip header
                    for row in reader:
                        if len(row) >= 2:
                            self.user_ids[row[0]] = row[1]
            else:
                # Create new user IDs file
                with open(self.user_ids_file, 'w', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow(['Name', 'UserID'])
                print(f"[INFO] Created new user IDs file: {self.user_ids_file}")
        except Exception as e:
            print(f"[ERROR] Failed to load user IDs: {e}")
    
    def _generate_user_id(self, name: str) -> str:
        """Generate a unique ID for a user based on their name"""
        # Use first 8 characters of MD5 hash for consistency
        hash_obj = hashlib.md5(name.encode())
        user_id = hash_obj.hexdigest()[:8].upper()
        return f"USR{user_id}"
    
    def _get_or_create_user_id(self, name: str) -> str:
        """Get existing user ID or create a new one"""
        if name not in self.user_ids:
            user_id = self._generate_user_id(name)
            self.user_ids[name] = user_id
            
            # Save to file
            try:
                with open(self.user_ids_file, 'a', newline='') as f:
                    writer = csv.writer(f)
                    writer.writerow([name, user_id])
            except Exception as e:
                print(f"[ERROR] Failed to save user ID: {e}")
        
        return self.user_ids[name]
    
    def _create_status_log_file(self):
        """Create a new status log CSV file with headers"""
        try:
            with open(self.status_log_file, 'w', newline='') as f:
                writer = csv.writer(f)
                writer.writerow(['Timestamp', 'Name', 'UserID', 'Status', 'Check_In_Time', 'Check_Out_Time', 'Duration'])
            print(f"[INFO] Created new status log file: {self.status_log_file}")
        except Exception as e:
            print(f"[ERROR] Failed to create status log file: {e}")
    
    def update_user_directory(self):
        """Create or update the User Directory sheet in Excel with all enrolled users"""
        try:
            if not EXCEL_AVAILABLE or not os.path.exists(self.attendance_file):
                return
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
                
                # Create or get User Directory sheet
                if 'User Directory' in wb.sheetnames:
                    ws = wb['User Directory']
                    # Clear existing data (keep header)
                    ws.delete_rows(2, ws.max_row)
                else:
                    ws = wb.create_sheet('User Directory', 0)  # Insert at beginning
                    
                    # Create header
                    ws.merge_cells('A1:E1')
                    ws['A1'] = 'USER DIRECTORY'
                    ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
                    ws['A1'].fill = PatternFill(start_color="2E86AB", end_color="2E86AB", fill_type="solid")
                    ws['A1'].alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Column headers
                    headers = ['#', 'Name', 'User ID', 'Enrolled Date', 'Total Attendance Days']
                    ws.append(headers)
                    
                    for cell in ws[2]:
                        cell.font = Font(bold=True, size=11)
                        cell.fill = PatternFill(start_color="A8DADC", end_color="A8DADC", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
                    
                    # Set column widths
                    ws.column_dimensions['A'].width = 6
                    ws.column_dimensions['B'].width = 20
                    ws.column_dimensions['C'].width = 15
                    ws.column_dimensions['D'].width = 15
                    ws.column_dimensions['E'].width = 22
                
                # Get all users from user_ids
                users = sorted(self.user_ids.items(), key=lambda x: x[0])
                
                # Add user data
                for idx, (name, user_id) in enumerate(users, 1):
                    # Try to find enrollment date from first attendance
                    enrolled_date = "N/A"
                    total_days = 0
                    
                    # Count total attendance days across all sheets
                    for sheet_name in wb.sheetnames:
                        if sheet_name.startswith(f"{name}_"):
                            # This is a user sheet
                            user_sheet = wb[sheet_name]
                            
                            # Count days with attendance (has First In value)
                            for row in user_sheet.iter_rows(min_row=3, values_only=True):
                                if len(row) > 2 and row[2]:  # Has First In time
                                    total_days += 1
                                    # Get earliest date as enrollment date
                                    if enrolled_date == "N/A" and row[0]:
                                        enrolled_date = row[0] if isinstance(row[0], str) else row[0].strftime('%Y-%m-%d')
                    
                    # If no attendance, use today as enrolled date
                    if enrolled_date == "N/A":
                        enrolled_date = datetime.now().strftime('%Y-%m-%d')
                    
                    row_data = [idx, name, user_id, enrolled_date, total_days]
                    ws.append(row_data)
                    
                    # Format the row
                    row_num = ws.max_row
                    ws.cell(row_num, 1).alignment = Alignment(horizontal="center")
                    ws.cell(row_num, 3).alignment = Alignment(horizontal="center")
                    ws.cell(row_num, 4).alignment = Alignment(horizontal="center")
                    ws.cell(row_num, 5).alignment = Alignment(horizontal="center")
                    
                    # Alternate row colors
                    if idx % 2 == 0:
                        for col in range(1, 6):
                            ws.cell(row_num, col).fill = PatternFill(start_color="F1FAEE", end_color="F1FAEE", fill_type="solid")
                
                # Add summary at bottom
                summary_row = ws.max_row + 2
                ws.merge_cells(f'A{summary_row}:B{summary_row}')
                ws[f'A{summary_row}'] = 'Total Users:'
                ws[f'A{summary_row}'].font = Font(bold=True)
                ws[f'A{summary_row}'].alignment = Alignment(horizontal="right")
                ws[f'C{summary_row}'] = len(users)
                ws[f'C{summary_row}'].font = Font(bold=True)
                ws[f'C{summary_row}'].fill = PatternFill(start_color="E63946", end_color="E63946", fill_type="solid")
                ws[f'C{summary_row}'].font = Font(bold=True, color="FFFFFF")
                ws[f'C{summary_row}'].alignment = Alignment(horizontal="center")
                
                wb.save(self.attendance_file)
                wb.close()
                
                print(f"[INFO] User Directory updated with {len(users)} users")
        
        except Exception as e:
            print(f"[ERROR] Failed to update user directory: {e}")
            import traceback
            traceback.print_exc()
    
    def _create_attendance_file(self):
        """Create a new attendance Excel file with monthly format - one sheet per user"""
        try:
            if not EXCEL_AVAILABLE:
                print("[ERROR] openpyxl not installed")
                return
            
            wb = Workbook()
            # Remove default sheet
            wb.remove(wb.active)
            
            # Create a template sheet (will be used when new users are added)
            ws = wb.create_sheet("Template")
            
            # Get current month/year for headers
            now = datetime.now()
            month_year = now.strftime('%B %Y')
            
            # Headers - Date column + Time In/Out columns for each day
            headers = ['Date', 'Day', 'First In', 'Last Out', 'Total Hours', 'Status', 'Late', 'OT']
            ws.append(headers)
            
            # Style headers
            for cell in ws[1]:
                cell.font = Font(bold=True, size=12, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal="center", vertical="center")
            
            # Set column widths
            ws.column_dimensions['A'].width = 12  # Date
            ws.column_dimensions['B'].width = 10  # Day
            ws.column_dimensions['C'].width = 10  # First In
            ws.column_dimensions['D'].width = 10  # Last Out
            ws.column_dimensions['E'].width = 12  # Total Hours
            ws.column_dimensions['F'].width = 12  # Status
            ws.column_dimensions['G'].width = 8   # Late
            ws.column_dimensions['H'].width = 8   # OT
            
            # Hide template sheet
            ws.sheet_state = 'hidden'
            
            wb.save(self.attendance_file)
            print(f"[INFO] Created new attendance file: {self.attendance_file}")
        except Exception as e:
            print(f"[ERROR] Failed to create attendance file: {e}")
    
    def _get_or_create_user_sheet(self, wb, name: str, current_date: datetime):
        """Get or create a sheet for a specific user with monthly calendar"""
        month_year = current_date.strftime('%B_%Y')
        sheet_name = f"{name}_{month_year}"
        
        # Check if sheet exists
        if sheet_name in wb.sheetnames:
            return wb[sheet_name]
        
        # Get or create user ID
        user_id = self._get_or_create_user_id(name)
        
        # Create new sheet for this user
        ws = wb.create_sheet(sheet_name)
        
        # Add header with user name, ID and month
        ws.merge_cells('A1:H1')
        ws['A1'] = f"{name} (ID: {user_id}) - {current_date.strftime('%B %Y')}"
        ws['A1'].font = Font(bold=True, size=14)
        ws['A1'].alignment = Alignment(horizontal="center")
        ws['A1'].fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        ws['A1'].font = Font(bold=True, size=14, color="FFFFFF")
        
        # Column headers
        headers = ['Date', 'Day', 'First In', 'Last Out', 'Total Hours', 'Status', 'Late', 'OT']
        ws.append(headers)
        
        for cell in ws[2]:
            cell.font = Font(bold=True, size=11)
            cell.fill = PatternFill(start_color="B4C7E7", end_color="B4C7E7", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center")
        
        # Set column widths
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 10
        ws.column_dimensions['C'].width = 10
        ws.column_dimensions['D'].width = 10
        ws.column_dimensions['E'].width = 12
        ws.column_dimensions['F'].width = 12
        ws.column_dimensions['G'].width = 8
        ws.column_dimensions['H'].width = 8
        
        # Pre-fill all days of the month
        import calendar
        year = current_date.year
        month = current_date.month
        days_in_month = calendar.monthrange(year, month)[1]
        
        for day in range(1, days_in_month + 1):
            date_obj = datetime(year, month, day)
            date_str = date_obj.strftime('%Y-%m-%d')
            day_name = date_obj.strftime('%a')
            
            row = [date_str, day_name, '', '', '', '', '', '']
            ws.append(row)
            
            # Color weekends differently
            row_num = ws.max_row
            if day_name in ['Sat', 'Sun']:
                for col in range(1, 9):
                    ws.cell(row_num, col).fill = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        
        # Add monthly summary section
        summary_row = ws.max_row + 2
        ws.merge_cells(f'A{summary_row}:H{summary_row}')
        ws[f'A{summary_row}'] = 'MONTHLY SUMMARY'
        ws[f'A{summary_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{summary_row}'].fill = PatternFill(start_color="E67E22", end_color="E67E22", fill_type="solid")
        ws[f'A{summary_row}'].alignment = Alignment(horizontal="center")
        
        summary_row += 1
        summary_labels = [
            ('A', 'Total Working Days:', 'B'),
            ('D', 'Total Hours Worked:', 'E'),
            ('A', 'Days Late:', 'B'),
            ('D', 'Total Late Time:', 'E'),
            ('A', 'Days with OT:', 'B'),
            ('D', 'Total OT:', 'E')
        ]
        
        for i in range(0, len(summary_labels), 2):
            label1_col, label1_text, value1_col = summary_labels[i]
            label2_col, label2_text, value2_col = summary_labels[i+1]
            
            ws[f'{label1_col}{summary_row}'] = label1_text
            ws[f'{label1_col}{summary_row}'].font = Font(bold=True)
            ws[f'{value1_col}{summary_row}'] = '=COUNTA(C3:C' + str(3 + days_in_month - 1) + ')' if 'Working Days' in label1_text else '0'
            
            ws[f'{label2_col}{summary_row}'] = label2_text
            ws[f'{label2_col}{summary_row}'].font = Font(bold=True)
            ws[f'{value2_col}{summary_row}'] = '0h 0m'
            
            summary_row += 1
        
        return ws
    
    def _load_today_checkins(self):
        """Load check-in times from today to enforce cooldown"""
        today = datetime.now().strftime('%Y-%m-%d')
        
        try:
            if not os.path.exists(self.attendance_file):
                return
            
            if not EXCEL_AVAILABLE:
                return
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
                ws = wb.active
            
                for row in ws.iter_rows(min_row=2, values_only=True):
                    if len(row) < 3:
                        continue
                    name = row[0]
                    day = row[1]
                    time_in = row[2]
                    
                    if day == today and name and time_in:
                        try:
                            timestamp = datetime.strptime(f"{today} {time_in}", '%Y-%m-%d %H:%M:%S')
                            if name not in self.last_checkin or timestamp > self.last_checkin[name]:
                                self.last_checkin[name] = timestamp
                        except:
                            pass

                wb.close()
        except Exception as e:
            print(f"[WARNING] Failed to load today's check-ins: {e}")
    
    def can_checkin(self, name: str) -> bool:
        """
        Check if person can check in (cooldown period)
        
        Args:
            name: Person's name
            
        Returns:
            True if can check in, False if in cooldown period
        """
        if name not in self.last_checkin:
            return True
        
        time_since_last = datetime.now() - self.last_checkin[name]
        cooldown = timedelta(seconds=config.CHECK_IN_COOLDOWN)
        
        return time_since_last >= cooldown
    
    def get_cooldown_remaining(self, name: str) -> int:
        """
        Get remaining cooldown time in seconds
        
        Args:
            name: Person's name
            
        Returns:
            Remaining seconds, or 0 if no cooldown
        """
        if name not in self.last_checkin:
            return 0
        
        time_since_last = datetime.now() - self.last_checkin[name]
        cooldown = timedelta(seconds=config.CHECK_IN_COOLDOWN)
        remaining = cooldown - time_since_last
        
        return max(0, int(remaining.total_seconds()))
    
    def record_event(self, name: str, event: str, confidence: float = 1.0) -> bool:
        """
        Record an attendance event to Excel - Monthly format with one sheet per user
        
        Args:
            name: Person's name
            event: Event type (CHECK_IN, CHECK_OUT, ACCESS_GRANTED, ACCESS_DENIED)
            confidence: Recognition confidence (0-1)
            
        Returns:
            True if successful, False otherwise
        """
        try:
            if not EXCEL_AVAILABLE:
                print("[ERROR] openpyxl not installed")
                return False
            
            now = datetime.now()
            date_str = now.strftime('%Y-%m-%d')
            time_str = now.strftime('%H:%M:%S')
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
                
                # Get or create user's monthly sheet
                ws = self._get_or_create_user_sheet(wb, name, now)
                
                # Find row for today (skip header rows 1 and 2)
                user_row = None
                for idx, row in enumerate(ws.iter_rows(min_row=3, values_only=False), start=3):
                    if row[0].value == date_str:
                        user_row = idx
                        break
                
                if not user_row:
                    print(f"[ERROR] Date {date_str} not found in {name}'s sheet")
                    wb.close()
                    return False
                
                if event == 'CHECK_IN':
                    # Update First In (only if empty - preserve first check-in)
                    if not ws.cell(user_row, 3).value:
                        ws.cell(user_row, 3, time_str)
                        
                        # Determine punctuality
                        cutoff_time = datetime.strptime("08:30:00", "%H:%M:%S").time()
                        if now.time() > cutoff_time:
                            ws.cell(user_row, 6, 'LATE')
                            ws.cell(user_row, 6).font = Font(bold=True, color="FF0000")
                            
                            # Calculate late time
                            cutoff_dt = datetime.strptime("08:30:00", "%H:%M:%S")
                            late_duration = now - cutoff_dt.replace(year=now.year, month=now.month, day=now.day)
                            late_minutes = int(late_duration.total_seconds() // 60)
                            late_hours = late_minutes // 60
                            late_mins = late_minutes % 60
                            if late_hours > 0:
                                time_late = f"{late_hours}h {late_mins}m"
                            else:
                                time_late = f"{late_mins}m"
                            ws.cell(user_row, 7, time_late)
                            ws.cell(user_row, 7).font = Font(bold=True, color="FF0000")
                        else:
                            ws.cell(user_row, 6, 'ON TIME')
                            ws.cell(user_row, 6).font = Font(bold=True, color="00B050")
                            ws.cell(user_row, 7, '0m')
                    else:
                        # Already has check-in, toggle by clearing checkout if exists
                        if ws.cell(user_row, 4).value:
                            ws.cell(user_row, 4).value = None  # Clear Last Out
                            ws.cell(user_row, 5).value = None  # Clear Total
                            ws.cell(user_row, 8).value = None  # Clear OT
                
                elif event == 'CHECK_OUT':
                    # Check if has check-in
                    if not ws.cell(user_row, 3).value:
                        print(f"[DEBUG] Refusing CHECK OUT - no check-in time for {date_str}")
                        wb.close()
                        return False
                    
                    # Always update Last Out (keep last checkout)
                    ws.cell(user_row, 4, time_str)
                    
                    # Calculate total time
                    time_in = ws.cell(user_row, 3).value
                    if time_in:
                        try:
                            in_dt = datetime.strptime(time_in, "%H:%M:%S")
                            out_dt = datetime.strptime(time_str, "%H:%M:%S")
                            duration = out_dt - in_dt
                            hours = duration.seconds // 3600
                            minutes = (duration.seconds % 3600) // 60
                            total_str = f"{hours}h {minutes}m"
                            ws.cell(user_row, 5, total_str)  # Total Hours
                            
                            # Calculate overtime (after 5PM = 17:00)
                            ot_cutoff = datetime.strptime("17:00:00", "%H:%M:%S").time()
                            if now.time() > ot_cutoff:
                                ot_start = datetime.strptime("17:00:00", "%H:%M:%S")
                                ot_start = ot_start.replace(year=now.year, month=now.month, day=now.day)
                                ot_duration = now - ot_start
                                ot_minutes = int(ot_duration.total_seconds() // 60)
                                ot_hours = ot_minutes // 60
                                ot_mins = ot_minutes % 60
                                if ot_hours > 0:
                                    ot_str = f"{ot_hours}h {ot_mins}m"
                                else:
                                    ot_str = f"{ot_mins}m"
                                ws.cell(user_row, 8, ot_str)
                                ws.cell(user_row, 8).font = Font(bold=True, color="FFA500")
                            else:
                                ws.cell(user_row, 8, '0m')
                        except Exception as calc_err:
                            print(f"[WARN] Failed to calculate time: {calc_err}")
                
                wb.save(self.attendance_file)
                wb.close()
            
            # Update monthly summary
            self.update_monthly_summary(name, now)
            
            # Update last check-in time
            if event in ['CHECK_IN', 'ACCESS_GRANTED']:
                self.last_checkin[name] = now
            
            print(f"[INFO] Recorded: {name} - {event} at {time_str}")
            return True
            
        except Exception as e:
            print(f"[ERROR] Failed to record event: {e}")
            import traceback
            traceback.print_exc()
            return False
    
    def get_today_attendance(self) -> List[Dict[str, str]]:
        """
        Get all attendance records for today (multi-sheet format)
        
        Returns:
            List of attendance records
        """
        today = datetime.now().strftime('%Y-%m-%d')
        records = []
        
        try:
            if not EXCEL_AVAILABLE or not os.path.exists(self.attendance_file):
                return records
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
                
                # Iterate through all sheets (except Template)
                for sheet_name in wb.sheetnames:
                    if sheet_name == 'Template':
                        continue
                    
                    ws = wb[sheet_name]
                    
                    # Extract user name from sheet name
                    name = sheet_name.split('_')[0] if '_' in sheet_name else sheet_name
                    
                    # Find today's row (skip header rows 1 and 2)
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        if len(row) < 6:
                            continue
                        
                        date_str = row[0]
                        
                        if date_str == today:
                            records.append({
                                'Name': name,
                                'Date': date_str,
                                'Day': row[1],
                                'Time In': row[2] if row[2] else '',
                                'Time Out': row[3] if row[3] else '',
                                'Total': row[4] if row[4] else '',
                                'Status': row[5] if row[5] else '',
                                'Time Late': row[6] if len(row) > 6 and row[6] else '0m',
                                'Time OT': row[7] if len(row) > 7 and row[7] else '0m'
                            })
                            break
                
                wb.close()
        
        except Exception as e:
            print(f"[ERROR] Failed to read attendance: {e}")
        
        return records
    
    def get_attendance_summary(self) -> Dict[str, int]:
        """
        Get summary of today's attendance
        
        Returns:
            Dictionary with counts of different events
        """
        records = self.get_today_attendance()
        summary = {
            'total_checkins': 0,
            'total_checkouts': 0,
            'total_access_granted': 0,
            'total_access_denied': 0,
            'unique_people': set()
        }
        
        for record in records:
            event = record['Event']
            name = record['Name']
            
            if event == 'CHECK_IN':
                summary['total_checkins'] += 1
                summary['unique_people'].add(name)
            elif event == 'CHECK_OUT':
                summary['total_checkouts'] += 1
            elif event == 'ACCESS_GRANTED':
                summary['total_access_granted'] += 1
                summary['unique_people'].add(name)
            elif event == 'ACCESS_DENIED':
                summary['total_access_denied'] += 1
        
        summary['unique_people'] = len(summary['unique_people'])
        
        return summary
    
    def export_report(self, output_file: str, start_date: Optional[str] = None, 
                     end_date: Optional[str] = None):
        """
        Export attendance report to a file
        
        Args:
            output_file: Output file path
            start_date: Start date (YYYY-MM-DD), optional
            end_date: End date (YYYY-MM-DD), optional
        """
        try:
            records = []
            
            with open(self.attendance_file, 'r') as f:
                reader = csv.DictReader(f)
                for row in reader:
                    date = row['Date']
                    
                    # Filter by date range if specified
                    if start_date and date < start_date:
                        continue
                    if end_date and date > end_date:
                        continue
                    
                    records.append(row)
            
            # Write filtered records
            with open(output_file, 'w', newline='') as f:
                if records:
                    writer = csv.DictWriter(f, fieldnames=records[0].keys())
                    writer.writeheader()
                    writer.writerows(records)
            
            print(f"[INFO] Exported {len(records)} records to {output_file}")
            
        except Exception as e:
            print(f"[ERROR] Failed to export report: {e}")
    
    def get_user_status(self) -> Dict[str, Dict]:
        """
        Get status of all users with their last check-in/out from Excel (multi-sheet format)
        
        Returns:
            Dictionary with user status information
        """
        today = datetime.now().strftime('%Y-%m-%d')
        user_status = {}
        
        try:
            if not EXCEL_AVAILABLE or not os.path.exists(self.attendance_file):
                return user_status
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
                
                # Iterate through all sheets (except Template)
                for sheet_name in wb.sheetnames:
                    if sheet_name == 'Template':
                        continue
                    
                    ws = wb[sheet_name]
                    
                    # Extract user name from sheet name (format: "Name_Month_Year")
                    name = sheet_name.split('_')[0] if '_' in sheet_name else sheet_name
                    
                    # Find today's row (skip header rows 1 and 2)
                    for row in ws.iter_rows(min_row=3, values_only=True):
                        if len(row) < 6:
                            continue
                        
                        date_str = row[0]  # Date
                        day_name = row[1]  # Day
                        time_in = row[2]   # First In
                        time_out = row[3]  # Last Out
                        total = row[4]     # Total Hours
                        status_text = row[5]  # Status
                        time_late = row[6] if len(row) > 6 else '0m'
                        time_ot = row[7] if len(row) > 7 else '0m'
                        
                        if date_str == today and name:
                            # Check if currently in OT (after 5PM and still checked in)
                            is_ot = False
                            if time_in and not time_out:
                                current_time = datetime.now().time()
                                ot_cutoff = datetime.strptime("17:00:00", "%H:%M:%S").time()
                                if current_time > ot_cutoff:
                                    is_ot = True
                            
                            user_status[name] = {
                                'last_event': 'CHECK_OUT' if time_out else 'CHECK_IN',
                                'last_time': time_out if time_out else time_in,
                                'status': 'OUT' if time_out else 'IN',
                                'check_in_time': time_in,
                                'check_out_time': time_out,
                                'duration': total,
                                'first_check_in': time_in,
                                'is_overtime': is_ot,
                                'time_ot': time_ot if time_ot else '0m'
                            }
                            break  # Found today's row for this user

                wb.close()
        
        except BadZipFile:
            print(f"[ERROR] Failed to get user status: attendance.xlsx is corrupted (not a zip file)")
        except Exception as e:
            print(f"[ERROR] Failed to get user status: {e}")
        
        return user_status
    
    def log_status_to_file(self):
        """Log current status of all users to file"""
        try:
            user_status = self.get_user_status()
            timestamp = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
            
            with open(self.status_log_file, 'a', newline='') as f:
                writer = csv.writer(f)
                for name, status in user_status.items():
                    user_id = self._get_or_create_user_id(name)
                    writer.writerow([
                        timestamp,
                        name,
                        user_id,
                        status['status'],
                        status.get('check_in_time', 'N/A'),
                        status.get('check_out_time', 'N/A'),
                        status.get('duration', 'N/A')
                    ])
            
            print(f"[INFO] Logged status for {len(user_status)} users")
            return True
        
        except Exception as e:
            print(f"[ERROR] Failed to log status: {e}")
            return False
    
    def create_monthly_report(self):
        """Create or update monthly report sheet"""
        try:
            if not EXCEL_AVAILABLE or not os.path.exists(self.attendance_file):
                return
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
            
            # Get or create Monthly Report sheet
            if 'Monthly Report' in wb.sheetnames:
                ws_monthly = wb['Monthly Report']
                ws_monthly.delete_rows(2, ws_monthly.max_row)  # Keep header, delete data
            else:
                ws_monthly = wb.create_sheet('Monthly Report', 1)
                # Create headers
                headers = ['Name', 'Days Late', 'OT Days', 'Total Time Late', 'Total Time OT']
                ws_monthly.append(headers)
                
                # Style headers
                for cell in ws_monthly[1]:
                    cell.font = Font(bold=True, size=12, color="FFFFFF")
                    cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                # Set column widths
                ws_monthly.column_dimensions['A'].width = 15
                ws_monthly.column_dimensions['B'].width = 12
                ws_monthly.column_dimensions['C'].width = 12
                ws_monthly.column_dimensions['D'].width = 15
                ws_monthly.column_dimensions['E'].width = 15
            
            # Get daily sheet
            ws_daily = wb.active
            
            # Calculate monthly statistics
            current_year_month = datetime.now().strftime('%Y-%m')
            user_stats = {}
            
            for row in ws_daily.iter_rows(min_row=2, values_only=True):
                if len(row) < 8:
                    continue
                
                name = row[0]
                day = row[1]
                status = row[5]  # Status column
                time_late = row[6]  # Time Late column
                time_ot = row[7]  # Time OT column
                
                if not name or not day:
                    continue
                
                # Check if date is in current month
                try:
                    day_str = day if isinstance(day, str) else day.strftime('%Y-%m-%d')
                    if not day_str.startswith(current_year_month):
                        continue
                except:
                    continue
                
                if name not in user_stats:
                    user_stats[name] = {
                        'days_late': 0,
                        'ot_days': 0,
                        'total_late_minutes': 0,
                        'total_ot_minutes': 0
                    }
                
                # Count late days
                if status == 'LATE':
                    user_stats[name]['days_late'] += 1
                    # Parse time late
                    if time_late and time_late != '0m':
                        user_stats[name]['total_late_minutes'] += self._parse_time_to_minutes(time_late)
                
                # Count OT days
                if time_ot and time_ot != '0m':
                    user_stats[name]['ot_days'] += 1
                    user_stats[name]['total_ot_minutes'] += self._parse_time_to_minutes(time_ot)
            
            # Write data to monthly report
            row_num = 2
            for name in sorted(user_stats.keys()):
                stats = user_stats[name]
                total_late = self._format_time_from_minutes(stats['total_late_minutes'])
                total_ot = self._format_time_from_minutes(stats['total_ot_minutes'])
                
                ws_monthly.append([
                    name,
                    stats['days_late'],
                    stats['ot_days'],
                    total_late,
                    total_ot
                ])
                
                # Style data rows
                for cell in ws_monthly[row_num]:
                    cell.alignment = Alignment(horizontal="center", vertical="center")
                
                row_num += 1
            
                wb.save(self.attendance_file)
                wb.close()
            print(f"[INFO] Monthly report updated")
        
        except BadZipFile:
            print(f"[ERROR] Failed to create monthly report: attendance.xlsx is corrupted (not a zip file)")
        except Exception as e:
            print(f"[ERROR] Failed to create monthly report: {e}")
    
    def _parse_time_to_minutes(self, time_str: str) -> int:
        """Parse time string like '1h 30m' or '30m' to total minutes"""
        try:
            time_str = time_str.strip()
            total_minutes = 0
            
            if 'h' in time_str:
                parts = time_str.split('h')
                hours = int(parts[0].strip())
                total_minutes += hours * 60
                
                if 'm' in parts[1]:
                    minutes = int(parts[1].replace('m', '').strip())
                    total_minutes += minutes
            elif 'm' in time_str:
                minutes = int(time_str.replace('m', '').strip())
                total_minutes += minutes
            
            return total_minutes
        except:
            return 0
    
    def _format_time_from_minutes(self, minutes: int) -> str:
        """Format minutes to 'Xh Ym' format"""
        if minutes == 0:
            return '0m'
        
        hours = minutes // 60
        mins = minutes % 60
        
        if hours > 0:
            return f"{hours}h {mins}m"
        else:
            return f"{mins}m"
    
    def update_monthly_summary(self, name: str, current_date: datetime = None):
        """Update monthly summary for a user's sheet"""
        if current_date is None:
            current_date = datetime.now()
        
        try:
            if not EXCEL_AVAILABLE or not os.path.exists(self.attendance_file):
                return
            
            with self._excel_lock:
                wb = load_workbook(self.attendance_file)
                month_year = current_date.strftime('%B_%Y')
                sheet_name = f"{name}_{month_year}"
                
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    return
                
                ws = wb[sheet_name]
                
                # Find summary section (starts after all days)
                year = current_date.year
                month = current_date.month
                days_in_month = calendar.monthrange(year, month)[1]
                data_end_row = 2 + days_in_month
                summary_start = data_end_row + 2
                
                # Calculate statistics
                total_days_worked = 0
                total_hours_minutes = 0
                days_late = 0
                total_late_minutes = 0
                days_with_ot = 0
                total_ot_minutes = 0
                
                for row_num in range(3, data_end_row + 1):
                    first_in = ws.cell(row_num, 3).value
                    total_hours = ws.cell(row_num, 5).value
                    time_late = ws.cell(row_num, 7).value
                    time_ot = ws.cell(row_num, 8).value
                    
                    # Count working days
                    if first_in:
                        total_days_worked += 1
                    
                    # Sum total hours
                    if total_hours:
                        total_hours_minutes += self._parse_time_to_minutes(total_hours)
                    
                    # Count late days
                    if time_late and time_late != '0m':
                        days_late += 1
                        total_late_minutes += self._parse_time_to_minutes(time_late)
                    
                    # Count OT days
                    if time_ot and time_ot != '0m':
                        days_with_ot += 1
                        total_ot_minutes += self._parse_time_to_minutes(time_ot)
                
                # Update summary cells
                summary_data_row = summary_start + 1
                ws.cell(summary_data_row, 2, total_days_worked)  # Total Working Days
                ws.cell(summary_data_row, 5, self._format_time_from_minutes(total_hours_minutes))  # Total Hours
                
                summary_data_row += 1
                ws.cell(summary_data_row, 2, days_late)  # Days Late
                ws.cell(summary_data_row, 5, self._format_time_from_minutes(total_late_minutes))  # Total Late Time
                
                summary_data_row += 1
                ws.cell(summary_data_row, 2, days_with_ot)  # Days with OT
                ws.cell(summary_data_row, 5, self._format_time_from_minutes(total_ot_minutes))  # Total OT
                
                wb.save(self.attendance_file)
                wb.close()
                
        except Exception as e:
            print(f"[ERROR] Failed to update monthly summary: {e}")
    
    def check_in_out(self, name: str) -> str:
        """
        Toggle check-in/check-out for a person
        
        Args:
            name: Person's name
            
        Returns:
            Status message: 'CHECKED IN', 'CHECKED OUT', or 'SUCCESS' (already in)
        """
        user_status = self.get_user_status()
        current_status = user_status.get(name, {}).get('status', 'NONE')
        
        print(f"[DEBUG] check_in_out called for {name}, current status: {current_status}")
        
        # Debug log to CSV
        debug_log = os.path.join(os.path.dirname(self.attendance_file), 'debug_log.csv')
        with open(debug_log, 'a') as f:
            from datetime import datetime
            now = datetime.now()
            f.write(f"{now.strftime('%H:%M:%S')},{name},{current_status},")

        # Original toggle behavior:
        # - If currently IN -> CHECK OUT
        # - Else -> CHECK IN
        if name in user_status and user_status[name].get('status') == 'IN':
            print(f"[DEBUG] {name} is IN, attempting CHECK OUT")
            wrote = self.record_event(name, 'CHECK_OUT')
            if not wrote:
                print(f"[DEBUG] CHECK OUT refused for {name}")
                with open(debug_log, 'a') as f:
                    f.write(f"CHECKOUT_REFUSED\n")
                return 'SUCCESS'
            self.log_status_to_file()
            self.create_monthly_report()
            print(f"[DEBUG] {name} CHECKED OUT successfully")
            with open(debug_log, 'a') as f:
                f.write(f"CHECKED_OUT\n")
            return 'CHECKED OUT'

        print(f"[DEBUG] {name} not IN, attempting CHECK IN")
        wrote = self.record_event(name, 'CHECK_IN')
        if not wrote:
            print(f"[DEBUG] CHECK IN refused for {name}")
            with open(debug_log, 'a') as f:
                f.write(f"CHECKIN_REFUSED\n")
            return 'SUCCESS'
        self.log_status_to_file()
        print(f"[DEBUG] {name} CHECKED IN successfully")
        with open(debug_log, 'a') as f:
            f.write(f"CHECKED_IN\n")
        return 'CHECKED IN'
    
    def manual_checkout(self, name: str) -> str:
        """
        Manually check out a user
        
        Args:
            name: Person's name
            
        Returns:
            Status message
        """
        user_status = self.get_user_status()
        
        if name in user_status and user_status[name]['status'] == 'IN':
            self.record_event(name, 'CHECK_OUT')
            self.log_status_to_file()
            self.create_monthly_report()
            return 'CHECKED OUT'
        else:
            return 'NOT CHECKED IN'


if __name__ == "__main__":
    """Test the attendance tracker"""
    print("[INFO] Testing attendance tracker...")
    
    tracker = AttendanceTracker()
    
    # Test recording events
    tracker.record_event("Test User", "CHECK_IN", 0.95)
    tracker.record_event("Test User", "CHECK_OUT", 0.92)
    
    # Get summary
    summary = tracker.get_attendance_summary()
    print(f"\n[INFO] Today's Summary:")
    print(f"  Check-ins: {summary['total_checkins']}")
    print(f"  Check-outs: {summary['total_checkouts']}")
    print(f"  Access Granted: {summary['total_access_granted']}")
    print(f"  Access Denied: {summary['total_access_denied']}")
    print(f"  Unique People: {summary['unique_people']}")
    
    # Show today's records
    records = tracker.get_today_attendance()
    print(f"\n[INFO] Today's Records ({len(records)} total):")
    for record in records[-5:]:  # Show last 5
        print(f"  {record['Name']} - {record['Event']} at {record['Time']}")

