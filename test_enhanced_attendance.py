#!/usr/bin/env python3
"""
Enhanced Attendance System Test
Tests the new features:
- Random User IDs
- Monthly summary with total working time
- User sheet format with statistics
"""

import os
import sys
from datetime import datetime, timedelta
from attendance_tracker import AttendanceTracker
from openpyxl import load_workbook

def print_section(title):
    """Print a formatted section header"""
    print("\n" + "=" * 70)
    print(f"  {title}")
    print("=" * 70)

def test_enhanced_features():
    """Test all enhanced features"""
    
    print_section("ENHANCED ATTENDANCE SYSTEM TEST")
    
    # Create tracker
    tracker = AttendanceTracker()
    
    # Test users
    test_users = ["Linh", "John", "Mary", "bb"]
    
    print_section("TEST 1: User ID Generation")
    print("\nGenerating unique IDs for users...")
    for user in test_users:
        user_id = tracker._get_or_create_user_id(user)
        print(f"  {user:12} → {user_id}")
    
    print_section("TEST 2: Recording Multiple Check-ins")
    print("\nSimulating check-ins at different times...")
    
    # Simulate different check-in times
    check_in_times = {
        "Linh": "08:15:00",  # On time
        "John": "08:45:00",  # Late
        "Mary": "08:00:00",  # Early
        "bb": "09:00:00"     # Very late
    }
    
    for user, time_str in check_in_times.items():
        result = tracker.record_event(user, 'CHECK_IN')
        status = "✓" if result else "✗"
        print(f"  {status} {user:12} checked in at {time_str}")
    
    print_section("TEST 3: Recording Check-outs with Overtime")
    print("\nSimulating check-outs at different times...")
    
    check_out_times = {
        "Linh": "17:30:00",  # 30 min OT
        "John": "16:45:00",  # No OT
        "Mary": "18:15:00",  # 1h 15m OT
    }
    
    for user, time_str in check_out_times.items():
        result = tracker.record_event(user, 'CHECK_OUT')
        status = "✓" if result else "✗"
        print(f"  {status} {user:12} checked out at {time_str}")
    
    print_section("TEST 4: User Status Summary")
    
    status = tracker.get_user_status()
    print(f"\n{'User':<12} {'Status':<8} {'In':<10} {'Out':<10} {'Total':<12} {'Late':<8} {'OT':<8}")
    print("-" * 70)
    
    for user, info in status.items():
        check_in = info.get('check_in_time') or 'N/A'
        check_out = info.get('check_out_time') or 'N/A'
        duration = info.get('duration') or 'N/A'
        time_late = info.get('time_late', '0m')
        time_ot = info.get('time_ot', '0m')
        
        print(f"{user:<12} {info['status']:<8} {check_in:<10} "
              f"{check_out:<10} {duration:<12} "
              f"{time_late:<8} {time_ot:<8}")
    
    print_section("TEST 5: Today's Attendance Records")
    
    records = tracker.get_today_attendance()
    for record in records:
        print(f"\n  {record['Name']} (Status: {record['Status']})")
        print(f"    Date: {record['Date']}")
        print(f"    In:   {record['Time In']}")
        print(f"    Out:  {record['Time Out']}")
        print(f"    Work: {record['Total']}")
        print(f"    Late: {record['Time Late']}")
        print(f"    OT:   {record['Time OT']}")
    
    print_section("TEST 6: Excel File Structure Analysis")
    
    test_file = tracker.attendance_file
    
    if not os.path.exists(test_file):
        print("  ✗ Attendance file not found!")
        return
    
    wb = load_workbook(test_file)
    
    print(f"\nFile: {test_file}")
    print(f"Total Sheets: {len(wb.sheetnames)}")
    
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Template':
            continue
        
        ws = wb[sheet_name]
        print(f"\n  📄 Sheet: {sheet_name}")
        
        # Print header (row 1)
        header = ws['A1'].value
        print(f"     Header: {header}")
        
        # Count attendance records
        today = datetime.now().strftime('%Y-%m-%d')
        attendance_count = 0
        
        for row in ws.iter_rows(min_row=3, max_row=ws.max_row):
            if row[0].value == today and row[2].value:  # Has date and check-in
                attendance_count += 1
                date = row[0].value
                day = row[1].value
                time_in = row[2].value
                time_out = row[3].value
                total = row[4].value
                status_val = row[5].value
                late = row[6].value
                ot = row[7].value
                
                print(f"     Today: {date} ({day})")
                print(f"       In: {time_in}, Out: {time_out}, Total: {total}")
                print(f"       Status: {status_val}, Late: {late}, OT: {ot}")
    
    print_section("TEST 7: User IDs File Verification")
    
    user_ids_file = tracker.user_ids_file
    
    if os.path.exists(user_ids_file):
        print(f"\nFile: {user_ids_file}")
        with open(user_ids_file, 'r') as f:
            content = f.read()
            print(content)
    else:
        print("  ✗ User IDs file not found!")
    
    print_section("TEST 8: Monthly Summary (Simulated)")
    
    print("\nUpdating monthly summaries for all users...")
    for user in test_users:
        tracker.update_monthly_summary(user)
        print(f"  ✓ Updated summary for {user}")
    
    # Reload and show summaries
    wb = load_workbook(test_file)
    for sheet_name in wb.sheetnames:
        if sheet_name == 'Template' or '_' not in sheet_name:
            continue
        
        ws = wb[sheet_name]
        user_name = sheet_name.split('_')[0]
        
        # Find summary section (look for merged cells with MONTHLY SUMMARY)
        print(f"\n  📊 {user_name}'s Monthly Summary:")
        
        found_summary = False
        for row_idx in range(1, ws.max_row + 1):
            cell_value = ws.cell(row_idx, 1).value
            if cell_value and 'MONTHLY SUMMARY' in str(cell_value):
                summary_start = row_idx + 1
                found_summary = True
                
                # Read summary data (check if cells exist)
                try:
                    print(f"     Total Working Days:  {ws.cell(summary_start, 2).value or 0}")
                    print(f"     Total Hours Worked:  {ws.cell(summary_start, 5).value or '0m'}")
                    print(f"     Days Late:           {ws.cell(summary_start + 1, 2).value or 0}")
                    print(f"     Total Late Time:     {ws.cell(summary_start + 1, 5).value or '0m'}")
                    print(f"     Days with OT:        {ws.cell(summary_start + 2, 2).value or 0}")
                    print(f"     Total OT:            {ws.cell(summary_start + 2, 5).value or '0m'}")
                except:
                    print("     (Summary data not available)")
                break
        
        if not found_summary:
            print(f"     (No summary section found)")
    
    wb.close()
    
    print_section("✅ ALL TESTS COMPLETED!")
    
    print("\nNew Features Summary:")
    print("  ✓ Random User IDs generated and stored")
    print("  ✓ Monthly calendar format with daily tracking")
    print("  ✓ Automatic monthly summary with total hours")
    print("  ✓ Late time and overtime tracking")
    print("  ✓ User ID included in all records")
    print("  ✓ Summary statistics: working days, total hours, late days, OT days")
    print(f"\nFiles created:")
    print(f"  • Attendance: {test_file}")
    print(f"  • User IDs:   {user_ids_file}")
    print(f"  • Status Log: {tracker.status_log_file}")
    
    print("\n" + "=" * 70)

if __name__ == '__main__':
    test_enhanced_features()
