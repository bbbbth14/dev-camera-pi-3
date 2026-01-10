#!/usr/bin/env python3
"""
Excel File Reader for Raspberry Pi 3
Read and display attendance data from Excel files
"""

import os
import sys
from datetime import datetime

try:
    from openpyxl import load_workbook
    EXCEL_AVAILABLE = True
except ImportError:
    EXCEL_AVAILABLE = False
    print("[ERROR] openpyxl not installed")
    print("Install with: pip3 install openpyxl --break-system-packages")
    sys.exit(1)

def read_attendance_file(file_path):
    """Read and display attendance Excel file"""
    
    if not os.path.exists(file_path):
        print(f"[ERROR] File not found: {file_path}")
        return
    
    try:
        print("="*80)
        print(f"Reading: {file_path}")
        print("="*80)
        print()
        
        # Load workbook
        wb = load_workbook(file_path)
        ws = wb.active
        
        print(f"Sheet name: {ws.title}")
        print(f"Total rows: {ws.max_row}")
        print(f"Total columns: {ws.max_column}")
        print()
        
        # Read headers
        headers = []
        for cell in ws[1]:
            headers.append(cell.value)
        
        # Display headers
        print("Columns:")
        for i, header in enumerate(headers, 1):
            print(f"  {i}. {header}")
        print()
        
        # Display data
        print("="*80)
        print("ATTENDANCE DATA")
        print("="*80)
        
        # Format header row
        header_line = " | ".join([f"{h:^12}" for h in headers if h])
        print(header_line)
        print("-" * len(header_line))
        
        # Read and display all rows
        row_count = 0
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:  # If name is not empty
                row_data = " | ".join([f"{str(cell):^12}" if cell else f"{'':^12}" 
                                      for cell in row[:len(headers)]])
                print(row_data)
                row_count += 1
        
        print("-" * len(header_line))
        print(f"\nTotal records: {row_count}")
        
        wb.close()
        
        # Summary statistics
        print()
        print("="*80)
        print("SUMMARY")
        print("="*80)
        
        # Count unique people
        wb = load_workbook(file_path)
        ws = wb.active
        
        unique_people = set()
        checked_in = []
        checked_out = []
        
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0]:
                name = row[0]
                unique_people.add(name)
                
                # Check status (Time Out column)
                time_out = row[3] if len(row) > 3 else None
                if time_out:
                    checked_out.append(name)
                else:
                    checked_in.append(name)
        
        print(f"Total people: {len(unique_people)}")
        print(f"Currently IN: {len(checked_in)}")
        print(f"Currently OUT: {len(checked_out)}")
        print()
        
        if checked_in:
            print("People currently IN:")
            for name in set(checked_in):
                print(f"  • {name}")
        
        wb.close()
        
    except Exception as e:
        print(f"[ERROR] Failed to read Excel file: {e}")
        import traceback
        traceback.print_exc()

def list_excel_files():
    """List all Excel files in data directory"""
    data_dir = "data"
    
    if not os.path.exists(data_dir):
        print(f"[ERROR] Directory not found: {data_dir}")
        return []
    
    excel_files = []
    for file in os.listdir(data_dir):
        if file.endswith('.xlsx') or file.endswith('.xls'):
            excel_files.append(os.path.join(data_dir, file))
    
    return excel_files

def main():
    """Main function"""
    print()
    print("="*80)
    print("EXCEL FILE READER FOR RASPBERRY PI 3")
    print("="*80)
    print()
    
    # List available Excel files
    excel_files = list_excel_files()
    
    if not excel_files:
        print("No Excel files found in 'data' directory")
        return
    
    print("Available Excel files:")
    for i, file in enumerate(excel_files, 1):
        size = os.path.getsize(file)
        modified = datetime.fromtimestamp(os.path.getmtime(file))
        print(f"  {i}. {file}")
        print(f"     Size: {size:,} bytes | Modified: {modified.strftime('%Y-%m-%d %H:%M:%S')}")
    
    print()
    
    # If command line argument provided, use it
    if len(sys.argv) > 1:
        file_path = sys.argv[1]
    else:
        # Ask user to select
        try:
            choice = input(f"Select file (1-{len(excel_files)}) [default: 1]: ").strip()
            if not choice:
                choice = "1"
            
            index = int(choice) - 1
            if 0 <= index < len(excel_files):
                file_path = excel_files[index]
            else:
                print("[ERROR] Invalid selection")
                return
        except ValueError:
            print("[ERROR] Invalid input")
            return
        except KeyboardInterrupt:
            print("\n\nCancelled")
            return
    
    print()
    read_attendance_file(file_path)
    print()

if __name__ == "__main__":
    main()
